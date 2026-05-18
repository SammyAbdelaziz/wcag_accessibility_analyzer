"""
XLSX WCAG Analyzer
Reads Excel workbooks using openpyxl to surface structural accessibility issues.

WCAG criteria covered:

  CONFIRMED (xml_direct):
    2.4.2  — workbook title absent or empty (document properties)
    3.1.1  — workbook language not set
    1.1.1  — charts without alt text / titles
    1.3.1  — merged cells used in data tables (disrupts table structure for AT)
    1.3.1  — data tables with no identifiable header row
    2.4.4  — hyperlinks with generic or URL-only display text

    POSSIBLE (xml_inferred):
        2.4.5  — large workbook lacks a table-of-contents sheet or named destinations

  POSSIBLE (xml_inferred):
    1.4.4  — text smaller than 8pt (tiny hardcoded font sizes)
    1.4.1  — data communicated by cell color alone (no pattern or text cue)
    1.3.1  — sheets with no visible column/row structure (freeform layout)
"""
from __future__ import annotations

import io
import re
from typing import List, Optional, Dict, Any, Tuple

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from wcag.models import (
    FactSheet, Finding,
    Severity, ConfidenceTier, EvidenceSource, CONFIDENCE_LABEL,
)
from wcag.common.safe_xml import SAFE_XML_PARSER
from wcag.common.safe_zip import open_safe_zip

GENERIC_LINK_TEXT = re.compile(
    r'^(click here|click|here|this link|learn more|more|read more|link|url|see here|https?://.*)$',
    re.IGNORECASE,
)
URL_PATTERN = re.compile(r'^https?://', re.IGNORECASE)


class XlsxAnalyzer:
    """Analyze an Excel workbook for WCAG 2.1 A/AA issues."""

    def __init__(self, file_bytes: bytes, filename: str):
        self.file_bytes = file_bytes
        self.filename = filename
        self.fact_sheet = FactSheet(
            filename=filename,
            file_type='xlsx',
        )

    def analyze(self) -> FactSheet:
        if not _OPENPYXL_AVAILABLE:
            self.fact_sheet.confirmed_findings.append(Finding(
                criterion_id="1.3.1",
                criterion_name="Info and Relationships",
                wcag_level="A",
                issue="openpyxl is not installed — XLSX analysis is unavailable.",
                evidence="Import of openpyxl failed.",
                severity=Severity.CRITICAL,
                why_it_matters="Cannot analyze XLSX without openpyxl.",
                remediation_steps=["Install openpyxl: pip install openpyxl>=3.1.2"],
                confidence_tier=ConfidenceTier.CONFIRMED,
                confidence_label="high",
                confidence_rationale="Import error is definitive.",
                evidence_source=EvidenceSource.XML_DIRECT,
                location="Analysis engine",
                remediation_id="xlsx_missing_dep",
            ))
            return self.fact_sheet

        try:
            # keep_vba=False strips any embedded VBA project before parse so the
            # in-memory workbook (and any downstream remediator output) cannot
            # re-emit attacker-supplied macros. data_only=True avoids formula
            # re-evaluation. keep_links=True is intentional for link-text rules.
            wb = load_workbook(
                io.BytesIO(self.file_bytes),
                data_only=True,
                keep_links=True,
                keep_vba=False,
            )
        except Exception as exc:
            self.fact_sheet.confirmed_findings.append(Finding(
                criterion_id="1.3.1",
                criterion_name="Info and Relationships",
                wcag_level="A",
                issue="Workbook could not be opened — file may be corrupt or password-protected.",
                evidence=str(exc)[:200],
                severity=Severity.CRITICAL,
                why_it_matters="A corrupt workbook cannot be analyzed or used by assistive technology.",
                remediation_steps=[
                    "Verify the file is a valid .xlsx (not .xls or .xlsm requiring macros).",
                    "Remove any password protection before analysis.",
                ],
                confidence_tier=ConfidenceTier.CONFIRMED,
                confidence_label="high",
                confidence_rationale="openpyxl raised an exception on open.",
                evidence_source=EvidenceSource.XML_DIRECT,
                location="Workbook root",
                remediation_id="xlsx_corrupt",
            ))
            return self.fact_sheet

        # Populate fact sheet metadata
        props = wb.properties
        self.fact_sheet.document_title = getattr(props, 'title', None) or None
        self.fact_sheet.document_language = getattr(props, 'language', None) or None
        self.fact_sheet.slide_count = len(wb.sheetnames)  # reuse slide_count for sheet count

        self._run_rules(wb)
        return self.fact_sheet

    # ── Rules engine ─────────────────────────────────────────────────────────

    def _run_rules(self, wb) -> None:
        self._rule_2_4_2_workbook_title()
        self._rule_2_4_5_multiple_ways(wb)
        self._rule_3_1_1_language()
        self._rule_1_1_1_charts(wb)
        self._rule_1_1_1_images(wb)
        for ws in wb.worksheets:
            loc = f"Sheet '{ws.title}'"
            self._rule_1_3_1_merged_cells(ws, loc)
            self._rule_1_3_1_header_row(ws, loc)
            self._rule_1_3_1_freeform_layout(ws, loc)
            self._rule_1_3_2_spreadsheet_sequence(ws, loc)
            self._rule_1_4_10_reflow(ws, loc)
            self._rule_2_4_4_link_text(ws, loc)
            self._rule_1_4_4_tiny_text(ws, loc)
            self._rule_1_4_1_color_only(ws, loc)
            self._rule_2_4_6_header_quality(ws, loc)  # Phase A
        self._rule_3_1_2_language_of_parts(wb)  # Phase A
        for ws in wb.worksheets:
            loc = f"Sheet '{ws.title}'"
            self._rule_1_4_11_non_text_contrast(ws, loc)  # Phase B
            self._rule_1_4_3_text_contrast(ws, loc)  # Phase C
            self._rule_1_3_3_sensory_characteristics(ws, loc)  # Phase I
        self._rule_4_1_2_form_controls()  # Phase K
        self._rule_2_4_2_generic_sheet_names(wb)  # Phase M-refinements R1
        for ws in wb.worksheets:
            loc = f"Sheet '{ws.title}'"
            self._rule_1_3_1_unfrozen_header(ws, loc)  # Phase M-refinements R2

    def _rule_1_3_3_sensory_characteristics(self, ws, loc: str) -> None:
        """WCAG 1.3.3 — flag cell text that references UI elements only by
        color/shape/position. Strict regex on string-typed cell values."""
        from wcag.common.sensory_characteristics import find_sensory_phrases
        items = []
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                v_stripped = v.strip()
                if not v_stripped:
                    continue
                items.append((cell.coordinate, v_stripped))
        if not items:
            return
        offenders = find_sensory_phrases(items)
        if not offenders:
            return
        sample = "; ".join(
            f"{o['kind']}-only @ {o['index']}: \"{o['snippet']}\""
            for o in offenders[:3]
        )
        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="1.3.3",
            criterion_name="Sensory Characteristics",
            wcag_level="A",
            issue=(
                f"{len(offenders)} cell(s) on {loc} reference UI elements only "
                "by color, shape, or position."
            ),
            evidence=f"Sensory-only references: {sample}.",
            severity=Severity.MODERATE,
            why_it_matters=(
                "Users who are blind, color-blind, or using screen readers cannot "
                "identify a control or sheet area by its visual appearance or position. "
                "Cell instructions must include the labeled name of the target."
            ),
            remediation_steps=[
                f"📍 WHERE TO FIX: {len(offenders)} cell(s) on {loc}.",
                "  • Add the named target: 'Use the Submit button (red)' rather than 'use the red button'.",
                "  • Replace position-only directions with named tabs/sections.",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label="high",
            confidence_rationale=f"String cell values on {loc} scanned for sensory-only phrase patterns.",
            evidence_source=EvidenceSource.TEXT_CONTENT,
            location=loc,
            remediation_id=f"xlsx_sensory_{ws.title[:20]}",
            remediation_data={"sheet": ws.title, "references": offenders},
        ))

    def _rule_4_1_2_form_controls(self):
        """WCAG 4.1.2 Name, Role, Value — Excel form controls (checkboxes,
        list boxes, dropdowns) live in xl/ctrlProps/*.xml. Each <formControlPr>
        element should have an `objectName` attribute that becomes the
        accessible name. Strict: flag controls with empty/missing objectName.
        """
        from lxml import etree as _lxml_etree
        try:
            with open_safe_zip(self.file_bytes) as zf:
                ctrl_files = [n for n in zf.namelist()
                              if n.startswith('xl/ctrlProps/') and n.endswith('.xml')]
                if not ctrl_files:
                    return
                offenders: List[Dict[str, Any]] = []
                for cf in ctrl_files:
                    try:
                        content = zf.read(cf)
                    except KeyError:
                        continue
                    try:
                        root = _lxml_etree.fromstring(content, SAFE_XML_PARSER)
                    except _lxml_etree.XMLSyntaxError:
                        continue
                    # Root is <formControlPr>. Read attributes (no namespace
                    # for these inner attrs in standard XLSX form controls).
                    name = (root.get('objectName') or '').strip()
                    alt = (root.get('altText') or '').strip()
                    obj_type = (root.get('objectType') or 'control')
                    if not name and not alt:
                        offenders.append({
                            "file": cf.split('/')[-1],
                            "object_type": obj_type,
                        })
            if not offenders:
                return
            sample = "; ".join(
                f"{o['file']} ({o['object_type']})" for o in offenders[:3]
            )
            self.fact_sheet.confirmed_findings.append(Finding(
                criterion_id="4.1.2",
                criterion_name="Name, Role, Value",
                wcag_level="A",
                issue=(
                    f"{len(offenders)} Excel form control(s) have no "
                    "accessible name (no objectName, no altText)."
                ),
                evidence=f"Form controls without a name: {sample}.",
                severity=Severity.SERIOUS,
                why_it_matters=(
                    "Excel form controls (checkboxes, dropdowns, list boxes) without "
                    "an objectName or alt-text are announced by screen readers as just "
                    "'check box' or 'combo box' — the user has no idea what they are "
                    "selecting. WCAG 4.1.2 requires every UI component to expose a name."
                ),
                remediation_steps=[
                    f"📍 WHERE TO FIX: {len(offenders)} form control(s) on the worksheet.",
                    "  • Right-click the control → Format Control → Alt Text → fill in Title and Description.",
                    "  • For Developer-tab ActiveX controls: select the control, in the Name Box give it a meaningful name.",
                    "  • Avoid generic names like 'CheckBox1'; prefer 'OptIn-Newsletter' or similar.",
                ],
                confidence_tier=ConfidenceTier.CONFIRMED,
                confidence_label="high",
                confidence_rationale="objectName/altText attributes parsed directly from xl/ctrlProps/*.xml entries.",
                evidence_source=EvidenceSource.XML_DIRECT,
                location=f"{len(offenders)} form control(s)",
                remediation_id="xlsx_form_control_names",
                remediation_data={"controls": offenders},
            ))
        except Exception:
            return  # ctrlProps not present or unreadable; not an error

    def _rule_2_4_2_generic_sheet_names(self, wb):
        """WCAG 2.4.2 (refinement) — flag sheets named with default placeholders
        ('Sheet1', 'Sheet 2', 'Tabelle1', etc.). Screen-reader users navigate
        a workbook by sheet name; generic names are unhelpful.
        """
        import re as _re
        generic_pat = _re.compile(
            r"^(?:sheet|sheet\s|tab|tabelle|hoja|feuille|lap)\s*\d+$",
            _re.IGNORECASE,
        )
        offenders = []
        for ws in wb.worksheets:
            name = (ws.title or '').strip()
            if generic_pat.match(name):
                offenders.append(name)
        if not offenders:
            return
        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="2.4.2",
            criterion_name="Page Titled",
            wcag_level="A",
            issue=(
                f"{len(offenders)} worksheet(s) use default generic names: "
                f"{', '.join(offenders[:5])}."
            ),
            evidence=f"Generic sheet names: {', '.join(offenders[:5])}.",
            severity=Severity.MODERATE,
            why_it_matters=(
                "Screen-reader users navigate a workbook by sheet tab. Default names "
                "like 'Sheet1', 'Sheet2' tell them nothing about the content and force "
                "them to open each sheet to figure out what it contains. A descriptive "
                "name (e.g. 'Q4 Revenue', 'Headcount by Region') doubles as a heading."
            ),
            remediation_steps=[
                f"📍 WHERE TO FIX: {len(offenders)} sheet tab(s).",
                "  • Right-click each tab → Rename → enter a 1-3 word descriptive name.",
                "  • Use the same name your team uses verbally to refer to that data.",
                "  • Keep names under 31 characters (Excel's tab limit).",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label="high",
            confidence_rationale="Sheet titles compared against locale-aware default placeholder pattern.",
            evidence_source=EvidenceSource.XML_DIRECT,
            location=f"{len(offenders)} sheet(s)",
            remediation_id="xlsx_generic_sheet_names",
            remediation_data={"sheets": offenders},
        ))

    def _rule_1_3_1_unfrozen_header(self, ws, loc: str):
        """WCAG 1.3.1 (refinement) — large data tables without frozen header
        rows lose their headers when scrolled. Header row exists (per Phase A
        check) but freeze_panes is empty/Top-left. Fires only when the sheet
        has > 20 data rows so we don't nag tiny tables.
        """
        # Count rows with at least one non-empty cell. Don't rely on
        # `ws.max_row` / `ws.calculate_dimension()` — earlier rules may have
        # accessed empty cells, which expands those bounds.
        content_rows = 0
        for row in ws.iter_rows(values_only=True):
            if any(c is not None and (not isinstance(c, str) or c.strip()) for c in row):
                content_rows += 1
                if content_rows >= 20:
                    break
        if content_rows < 20:
            return
        # Only act on sheets with text in row 1 — likely a header.
        first_row_has_text = any(
            cell.value is not None and isinstance(cell.value, str) and cell.value.strip()
            for cell in next(ws.iter_rows(min_row=1, max_row=1), [])
        )
        if not first_row_has_text:
            return
        fp = ws.freeze_panes
        # freeze_panes is None or 'A1' or '' when nothing is frozen, or e.g.
        # 'A2' when row 1 is frozen, 'B2' when row1+colA frozen, etc.
        if fp and fp not in ('A1', 'A', '', None):
            # Check the row index in the freeze coord — if it's >= 2, row 1
            # IS frozen and we're done.
            import re as _re
            m = _re.match(r"^[A-Z]+(\d+)$", str(fp))
            if m and int(m.group(1)) >= 2:
                return  # header is frozen
        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="1.3.1",
            criterion_name="Info and Relationships",
            wcag_level="A",
            issue=(
                f"{loc} has at least 20 rows of data but the header row is not frozen — "
                "headers disappear when the user scrolls."
            ),
            evidence=f"freeze_panes={ws.freeze_panes!r}; content_rows>=20.",
            severity=Severity.MODERATE,
            why_it_matters=(
                "When a data table is taller than the visible area and the header row is "
                "not frozen, screen-reader users (and sighted users) lose the column "
                "context as soon as they scroll. WCAG 1.3.1 requires that the relationship "
                "between data and its header be programmatically determinable; freezing "
                "the header row preserves that relationship visually and via the "
                "TitleRegion accessibility API."
            ),
            remediation_steps=[
                f"📍 WHERE TO FIX: {loc}.",
                "  • View → Freeze Panes → Freeze Top Row.",
                "  • Or, click cell A2 and use View → Freeze Panes → Freeze Panes.",
                "  • Confirm by scrolling — row 1 should remain visible at the top.",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label="high",
            confidence_rationale=f"freeze_panes is {ws.freeze_panes!r}; row 1 is text but not frozen.",
            evidence_source=EvidenceSource.XML_DIRECT,
            location=loc,
            remediation_id=f"xlsx_unfrozen_header_{ws.title[:20]}",
            remediation_data={"sheet": ws.title, "freeze_panes": str(ws.freeze_panes)},
        ))

    # ── 2.4.2 Page Titled ────────────────────────────────────────────────────

    def _rule_2_4_2_workbook_title(self):
        """Check that the workbook has a non-empty document title in properties."""
        title = self.fact_sheet.document_title
        if title and title.strip():
            return
        import os, re as _re
        base = os.path.splitext(os.path.basename(self.filename))[0]
        # Humanize filename: replace separators, strip leading numbers+dash
        suggested = _re.sub(r'[-_]+', ' ', base).strip()
        suggested = _re.sub(r'^[\d\s]+', '', suggested).strip().title() or base

        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="2.4.2",
            criterion_name="Page Titled",
            wcag_level="A",
            issue="Workbook has no document title set in file properties.",
            evidence=f"Document properties Title field is {'empty' if title == '' else 'absent'}.",
            severity=Severity.MODERATE,
            why_it_matters=(
                "Screen readers announce the document title when a file is opened. "
                "Without one, users hear only the filename — which may be a UUID or "
                "internal reference code that conveys no useful information."
            ),
            remediation_steps=[
                "📍 WHERE TO FIX: File → Info → Properties (right panel) → Title field.",
                "",
                "HOW TO FIX:",
                "  • In Excel: File → Info → Properties → Title → type a descriptive title.",
                "  • Title should describe the workbook's purpose, e.g. 'Q4 2026 Budget Summary'.",
                f"  • Suggested title based on filename: '{suggested}'.",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label="high",
            confidence_rationale="Title field read directly from workbook document properties.",
            evidence_source=EvidenceSource.XML_DIRECT,
            location="File → Info → Properties → Title",
            remediation_id="xlsx_doc_title",
            remediation_data={"action": "set_workbook_title", "suggested_title": suggested},
        ))

    def _rule_2_4_5_multiple_ways(self, wb) -> None:
        """WCAG 2.4.5 — large workbooks should offer more than sheet-tab
        navigation alone. Heuristic: visible multi-sheet workbooks should have
        either a TOC-like sheet or user-defined names that act as destinations.
        """
        visible_sheets = [ws for ws in wb.worksheets if getattr(ws, 'sheet_state', 'visible') == 'visible']
        if len(visible_sheets) < 5:
            return

        toc_title_pattern = re.compile(
            r"^(contents?|table of contents|toc|index|overview|navigation|read ?me)$",
            re.IGNORECASE,
        )
        visible_titles = [str(ws.title or '').strip() for ws in visible_sheets]
        lower_titles = {title.lower() for title in visible_titles if title}

        toc_sheets: List[str] = []
        for ws in visible_sheets:
            sheet_title = str(ws.title or '').strip()
            if not toc_title_pattern.match(sheet_title):
                continue
            referenced_titles = set()
            max_rows = min(max(ws.max_row or 1, 1), 50)
            max_cols = min(max(ws.max_column or 1, 1), 8)
            for row in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols):
                for cell in row:
                    value = cell.value
                    if isinstance(value, str):
                        value_lower = value.strip().lower()
                        if not value_lower:
                            continue
                        for title in lower_titles:
                            if title != sheet_title.lower() and title in value_lower:
                                referenced_titles.add(title)
                    hyperlink = getattr(cell, 'hyperlink', None)
                    if hyperlink is not None:
                        target = str(getattr(hyperlink, 'target', '') or '').lower()
                        location = str(getattr(hyperlink, 'location', '') or '').lower()
                        link_blob = f"{target} {location}"
                        for title in lower_titles:
                            if title != sheet_title.lower() and title in link_blob:
                                referenced_titles.add(title)
            if len(referenced_titles) >= 2:
                toc_sheets.append(sheet_title)

        user_defined_names = []
        for name, defined_name in getattr(wb, 'defined_names', {}).items():
            clean_name = str(name or '').strip()
            if not clean_name or clean_name.lower().startswith('_xlnm.'):
                continue
            attr_text = str(getattr(defined_name, 'attr_text', '') or '').strip()
            if attr_text:
                user_defined_names.append(clean_name)

        if toc_sheets or user_defined_names:
            return

        sample_sheets = ', '.join(visible_titles[:5])
        self.fact_sheet.possible_findings.append(Finding(
            criterion_id="2.4.5",
            criterion_name="Multiple Ways",
            wcag_level="AA",
            issue=(
                f"Workbook has {len(visible_sheets)} visible sheets but no detectable table-of-contents sheet "
                "or named destinations for alternate navigation."
            ),
            evidence=(
                f"Visible sheets: {sample_sheets}. No TOC-like sheet with cross-sheet references and no user-defined names were found."
            ),
            severity=Severity.MINOR,
            why_it_matters=(
                "Large workbooks force keyboard and screen-reader users to move sheet-by-sheet when there is no index or named destination list. "
                "A contents sheet or named ranges provides a second way to locate content quickly."
            ),
            remediation_steps=[
                "📍 WHERE TO FIX: workbook-level navigation.",
                "  • Add a first sheet named 'Contents' or 'Index' listing the major sheets.",
                "  • Include links or clear references to the destination sheets from that index sheet.",
                "  • Or add meaningful named ranges so users can jump via the Name Box / Go To dialog.",
            ],
            confidence_tier=ConfidenceTier.POSSIBLE,
            confidence_label=CONFIDENCE_LABEL[EvidenceSource.XML_INFERRED],
            confidence_rationale=(
                "Workbook structure was inspected directly, but adequacy of navigation aids is heuristic: the workbook has many visible sheets and no TOC-like sheet or named destinations."
            ),
            evidence_source=EvidenceSource.XML_INFERRED,
            location="Workbook navigation",
            remediation_id="xlsx_multiple_ways",
            remediation_data={
                "visible_sheets": visible_titles,
                "toc_sheets": toc_sheets,
                "defined_names": user_defined_names[:10],
            },
        ))

    # ── 3.1.1 Language of Page ───────────────────────────────────────────────

    def _rule_3_1_1_language(self):
        """Check that the workbook has a language declared in document properties."""
        lang = self.fact_sheet.document_language
        if lang and lang.strip():
            return

        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="3.1.1",
            criterion_name="Language of Page",
            wcag_level="A",
            issue="Workbook document language is not set.",
            evidence="Document properties Language field is absent or empty.",
            severity=Severity.MODERATE,
            why_it_matters=(
                "Screen readers select the text-to-speech engine based on the declared "
                "language. Without it, content may be mispronounced using wrong phonetics."
            ),
            remediation_steps=[
                "📍 WHERE TO FIX: File → Info → Properties → Language, or via Review → Language.",
                "",
                "HOW TO FIX:",
                "  • Excel: File → Options → Language → set the Office display and proofing language.",
                "  • Review → Language → Set Proofing Language → select the primary language.",
                "  • Common values: 'en-US' (English US), 'en-GB' (English UK), 'fr-FR' (French).",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label="high",
            confidence_rationale="Language field read directly from workbook document properties.",
            evidence_source=EvidenceSource.XML_DIRECT,
            location="File → Info → Properties → Language",
            remediation_id="xlsx_doc_language",
            remediation_data={"action": "set_workbook_language", "suggested_lang": "en-US"},
        ))

    # ── 1.1.1 Non-text Content — Charts ──────────────────────────────────────

    def _rule_1_1_1_charts(self, wb) -> None:
        """Check that charts have a title (used as their accessible name by AT)."""
        charts_without_title: List[Dict] = []
        charts_with_title: int = 0

        for ws in wb.worksheets:
            for chart in ws._charts:
                loc = f"Sheet '{ws.title}'"
                try:
                    title_text = None
                    if chart.title is not None:
                        # chart.title can be a string or a RichText object
                        t = chart.title
                        if hasattr(t, 'text'):
                            title_text = str(t.text).strip()
                        elif hasattr(t, 'tx') and t.tx is not None:
                            title_text = str(t.tx).strip()
                        else:
                            title_text = str(t).strip() if t else ''
                    if title_text:
                        charts_with_title += 1
                    else:
                        charts_without_title.append({
                            'sheet': ws.title,
                            'type': type(chart).__name__,
                        })
                except Exception:
                    charts_without_title.append({'sheet': ws.title, 'type': 'Unknown'})

        if not charts_without_title:
            return

        sheet_names = list(dict.fromkeys(c['sheet'] for c in charts_without_title))
        examples = ', '.join(
            f"{c['type']} on sheet '{c['sheet']}'" for c in charts_without_title[:4]
        )
        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="1.1.1",
            criterion_name="Non-text Content",
            wcag_level="A",
            issue=(
                f"{len(charts_without_title)} chart(s) across {len(sheet_names)} sheet(s) "
                "have no title — screen readers cannot identify what they represent."
            ),
            evidence=f"Charts without title: {examples}.",
            severity=Severity.CRITICAL,
            why_it_matters=(
                "Charts without titles are announced as 'image' or 'chart' by screen readers "
                "with no indication of what data they display. This is a complete information "
                "barrier for blind or low-vision users."
            ),
            remediation_steps=[
                f"📍 WHERE TO FIX: Sheet(s): {', '.join(sheet_names[:5])}.",
                "",
                "HOW TO FIX:",
                "  • Click the chart → Chart Design tab → Add Chart Element → Chart Title → Above Chart.",
                "  • Type a descriptive title that explains what the chart shows, e.g. 'Monthly Revenue by Region Q4 2026'.",
                "  • Avoid generic titles like 'Chart 1' or 'Sales Data' — be specific about the data and time period.",
                "  • Also add alt text: right-click chart → Format Chart Area → Alt Text tab.",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label="high",
            confidence_rationale=(
                f"Chart title property read directly from workbook XML. "
                f"{len(charts_without_title)} chart(s) have absent or empty title."
            ),
            evidence_source=EvidenceSource.XML_DIRECT,
            location=f"Sheet(s): {', '.join(sheet_names[:5])}",
            remediation_id="xlsx_chart_titles",
            remediation_data={
                "action": "add_chart_titles",
                "charts": charts_without_title[:10],
            },
        ))

    # ── 1.3.1 Info and Relationships — Merged Cells ───────────────────────────

    def _rule_1_3_1_merged_cells(self, ws, loc: str) -> None:
        """Merged cells in data tables disrupt AT table navigation.

        Screen readers rely on row/column indices to announce 'Row 2, Column 3'.
        Merged cells shift indices unpredictably, breaking this navigation.
        This rule flags sheets that use merged cells in what appears to be a
        data table (has header-like row + data rows below it).
        """
        if not ws.merged_cells.ranges:
            return

        # Only flag if the sheet looks like a data table (has content in rows 1-3)
        has_data = False
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if any(cell is not None for cell in row):
                has_data = True
                break
        if not has_data:
            return

        merge_ranges = [str(r) for r in ws.merged_cells.ranges]
        merge_count = len(merge_ranges)
        examples = ', '.join(merge_ranges[:6])
        if merge_count > 6:
            examples += f' (and {merge_count - 6} more)'

        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="1.3.1",
            criterion_name="Info and Relationships",
            wcag_level="A",
            issue=(
                f"{loc}: {merge_count} merged cell range(s) detected — "
                "merging disrupts screen reader table navigation."
            ),
            evidence=f"Merged ranges: {examples}.",
            severity=Severity.SERIOUS,
            why_it_matters=(
                "Screen readers navigate tables by row and column. Merged cells cause "
                "the AT to announce wrong row/column positions and may skip cells entirely, "
                "making data tables incomprehensible."
            ),
            remediation_steps=[
                f"📍 WHERE TO FIX: {loc} — merged cell ranges: {', '.join(merge_ranges[:4])}.",
                "",
                "HOW TO FIX:",
                "  • Unmerge cells: select the merged range → Home → Merge & Center (toggle off).",
                "  • If merging is used for visual layout (spanning headers), use 'Center Across Selection' instead.",
                "    (Format Cells → Alignment → Horizontal → Center Across Selection).",
                "  • For table headers spanning columns, repeat the header text in each column instead of merging.",
                "  • If merging is decorative (title row), place the title above the table range as a separate row.",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label="high",
            confidence_rationale=(
                f"Merged cell ranges read directly from worksheet XML. "
                f"{merge_count} ranges found in {loc}."
            ),
            evidence_source=EvidenceSource.XML_DIRECT,
            location=loc,
            remediation_id=f"xlsx_merged_cells_{ws.title[:20]}",
            remediation_data={
                "action": "unmerge_cells",
                "sheet": ws.title,
                "merge_ranges": merge_ranges[:20],
            },
        ))

    # ── 1.3.1 Info and Relationships — Header Row ────────────────────────────

    def _rule_1_3_1_header_row(self, ws, loc: str) -> None:
        """Check that data tables have identifiable header rows.

        A data table without a header row means AT users cannot determine what
        each column represents. We look for sheets with ≥ 3 rows and ≥ 2 columns
        of data, and check whether the first row looks like headers (bold,
        distinct from data rows, or explicitly defined as a table header).

        If the sheet has no data, skip. If it has a defined Table object with
        headers, pass. Otherwise check for bold/distinguishable first row.
        """
        # Check if sheet has openpyxl Table objects with header rows defined
        if hasattr(ws, 'tables') and ws.tables:
            for tbl in ws.tables.values():
                # openpyxl Table has headerRowCount; if 0, no header
                if hasattr(tbl, 'headerRowCount') and tbl.headerRowCount == 0:
                    self.fact_sheet.confirmed_findings.append(Finding(
                        criterion_id="1.3.1",
                        criterion_name="Info and Relationships",
                        wcag_level="A",
                        issue=(
                            f"{loc}: Table '{tbl.displayName}' has headerRowCount=0 — "
                            "no header row is defined for this data table."
                        ),
                        evidence=f"Table '{tbl.displayName}' at range {tbl.ref} has no header row (headerRowCount=0).",
                        severity=Severity.SERIOUS,
                        why_it_matters=(
                            "Without a defined header row, screen readers cannot announce column names "
                            "as users navigate table cells. Users must memorise column positions manually."
                        ),
                        remediation_steps=[
                            f"📍 WHERE TO FIX: {loc} → Table '{tbl.displayName}' ({tbl.ref}).",
                            "  • Click in the table → Table Design tab → check 'Header Row'.",
                            "  • Ensure the first row of the table contains descriptive column labels.",
                        ],
                        confidence_tier=ConfidenceTier.CONFIRMED,
                        confidence_label="high",
                        confidence_rationale="headerRowCount=0 read directly from table XML definition.",
                        evidence_source=EvidenceSource.XML_DIRECT,
                        location=loc,
                        remediation_id=f"xlsx_table_header_{ws.title[:20]}_{tbl.displayName[:10]}",
                        remediation_data={
                            "action": "enable_header_row",
                            "sheet": ws.title,
                            "table": tbl.displayName,
                            "range": tbl.ref,
                        },
                    ))
            return  # Tables defined — header presence is handled above

        # Heuristic: look for a data block with ≥ 3 rows and ≥ 2 columns
        # where the first row is NOT bold/distinguished
        try:
            rows = list(ws.iter_rows(min_row=1, max_row=10))
            if len(rows) < 3:
                return
            # Count populated cells in row 1 and row 2
            row1 = rows[0]
            row2 = rows[1]
            populated_r1 = [c for c in row1 if c.value is not None]
            populated_r2 = [c for c in row2 if c.value is not None]
            if len(populated_r1) < 2 or len(populated_r2) < 2:
                return

            # Check if row 1 appears to be bold (common header indicator)
            bold_in_r1 = sum(1 for c in populated_r1 if c.font and c.font.bold)
            bold_in_r2 = sum(1 for c in populated_r2 if c.font and c.font.bold)

            # If an AutoFilter starts on row 1, treat row 1 as a declared
            # header surface even when formatting is plain.
            auto_filter_ref = getattr(ws.auto_filter, "ref", None)
            if auto_filter_ref:
                return

            # Infer likely headers from content semantics to avoid false
            # positives on clean but unformatted tables:
            # - first row has mostly text labels
            # - second row has at least one numeric/date-ish value
            # - first-row labels are unique enough to act as headers
            r1_values = [c.value for c in populated_r1]
            r2_values = [c.value for c in populated_r2]
            r1_text_like = [v for v in r1_values if isinstance(v, str) and v.strip()]
            r2_non_text = [v for v in r2_values if isinstance(v, (int, float))]
            r1_unique = len({str(v).strip().lower() for v in r1_text_like}) == len(r1_text_like)
            if len(r1_text_like) >= max(2, len(populated_r1) - 1) and r1_unique and len(r2_non_text) >= 1:
                return

            # If neither row has any bold text, flag as possible missing header
            if bold_in_r1 == 0 and bold_in_r2 == 0:
                col_count = len(populated_r1)
                row_count = ws.max_row or 0
                if row_count < 4:
                    return  # Too small to be a data table

                header_values = [str(c.value)[:20] for c in populated_r1[:5]]
                self.fact_sheet.possible_findings.append(Finding(
                    criterion_id="1.3.1",
                    criterion_name="Info and Relationships",
                    wcag_level="A",
                    issue=(
                        f"{loc}: Data table with {row_count} rows and {col_count} columns "
                        "may be missing a styled header row — column roles are unclear to AT."
                    ),
                    evidence=(
                        f"First row values: {header_values}. "
                        "No bold formatting or defined Table header found."
                    ),
                    severity=Severity.MODERATE,
                    why_it_matters=(
                        "Screen readers rely on table headers to announce column context. "
                        "Without clear headers, data cells are announced without any column label."
                    ),
                    remediation_steps=[
                        f"📍 WHERE TO CHECK: {loc}, row 1.",
                        "  • Bold the first row: select row 1 → Ctrl+B.",
                        "  • Better: format as a table: select data range → Insert → Table → check 'My table has headers'.",
                        "  • Descriptive column headers (e.g. 'Employee Name', 'Department') are more useful than abbreviations.",
                    ],
                    confidence_tier=ConfidenceTier.POSSIBLE,
                    confidence_label="medium",
                    confidence_rationale="No bold or formal table header detected; may be data without header or unformatted header.",
                    evidence_source=EvidenceSource.XML_INFERRED,
                    location=loc,
                    remediation_id=f"xlsx_header_row_{ws.title[:20]}",
                    remediation_data={"action": "add_header_row", "sheet": ws.title},
                ))
        except Exception:
            pass

    # ── 2.4.4 Link Purpose ───────────────────────────────────────────────────

    def _rule_2_4_4_link_text(self, ws, loc: str) -> None:
        """Check hyperlinks in worksheet for non-descriptive display text."""
        bad_links: List[Dict] = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink is None:
                    continue
                href = cell.hyperlink.target or ''
                display = str(cell.value or '').strip() if cell.value else ''

                is_raw_url = bool(URL_PATTERN.match(display))
                if not display or GENERIC_LINK_TEXT.match(display) or is_raw_url:
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    bad_links.append({
                        'cell': cell_ref,
                        'display': display or '(empty)',
                        'href': href[:80],
                    })

        if not bad_links:
            return

        examples = '; '.join(
            f"Cell {l['cell']}: '{l['display']}' → {l['href'][:40]}"
            for l in bad_links[:4]
        )
        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="2.4.4",
            criterion_name="Link Purpose (In Context)",
            wcag_level="A",
            issue=(
                f"{loc}: {len(bad_links)} hyperlink(s) have generic or empty display text "
                "that doesn't describe their destination."
            ),
            evidence=f"Links: {examples}.",
            severity=Severity.MODERATE,
            why_it_matters=(
                "Screen reader users often navigate by listing all links on a page. "
                "Generic text like 'Click here' or raw URLs give no indication of "
                "where the link leads."
            ),
            remediation_steps=[
                f"📍 WHERE TO FIX: {loc} — cells: {', '.join(l['cell'] for l in bad_links[:6])}.",
                "",
                "HOW TO FIX:",
                "  • Click each cell → right-click → Edit Hyperlink → change 'Text to display' to descriptive text.",
                "  • Good: 'Q4 2026 Budget Report (PDF)' | Bad: 'click here', 'link', raw URL.",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label="high",
            confidence_rationale="Hyperlink display text read directly from cell value and hyperlink XML.",
            evidence_source=EvidenceSource.XML_DIRECT,
            location=loc,
            remediation_id=f"xlsx_link_text_{ws.title[:20]}",
            remediation_data={
                "action": "fix_link_text",
                "sheet": ws.title,
                "links": bad_links[:10],
            },
        ))

    # ── 1.4.4 Resize Text — Tiny Font Sizes ──────────────────────────────────

    _TINY_FONT_PT = 8.0

    def _rule_1_4_4_tiny_text(self, ws, loc: str) -> None:
        """Flag cells with extremely small font sizes (< 8pt)."""
        tiny: List[Dict] = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if cell.font and cell.font.size and cell.font.size < self._TINY_FONT_PT:
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    tiny.append({
                        'cell': cell_ref,
                        'pt': cell.font.size,
                        'text': str(cell.value)[:40],
                    })

        if not tiny:
            return

        examples = '; '.join(
            f"Cell {t['cell']} ({t['pt']}pt): '{t['text']}'" for t in tiny[:4]
        )
        self.fact_sheet.possible_findings.append(Finding(
            criterion_id="1.4.4",
            criterion_name="Resize Text",
            wcag_level="AA",
            issue=(
                f"{loc}: {len(tiny)} cell(s) use font size below {self._TINY_FONT_PT}pt — "
                "unreadable for many users."
            ),
            evidence=f"Cells with tiny font: {examples}.",
            severity=Severity.MODERATE,
            why_it_matters=(
                "Text below 8pt is unreadable for most users, including those with normal vision "
                "at normal viewing distances. Excel's zoom does not always carry through to printed "
                "or exported forms."
            ),
            remediation_steps=[
                f"📍 WHERE TO FIX: {loc} — cells: {', '.join(t['cell'] for t in tiny[:6])}.",
                "  • Select the cell(s) → Home → Font Size → increase to at least 10pt.",
                "  • Body text: 11–12pt minimum; footnotes/annotations: 9–10pt minimum.",
            ],
            confidence_tier=ConfidenceTier.POSSIBLE,
            confidence_label="medium",
            confidence_rationale=f"Font size read from cell font property; {len(tiny)} cell(s) below {self._TINY_FONT_PT}pt.",
            evidence_source=EvidenceSource.XML_INFERRED,
            location=loc,
            remediation_id=f"xlsx_tiny_text_{ws.title[:20]}",
            remediation_data={"action": "increase_font_size", "sheet": ws.title, "cells": tiny[:10]},
        ))

    # ── 1.4.1 Use of Color ───────────────────────────────────────────────────

    def _rule_1_4_1_color_only(self, ws, loc: str) -> None:
        """Flag cells where background color appears to encode meaning without text cue.

        Heuristic: detect rows where adjacent cells have distinct fill colors but
        share the same text content pattern (e.g. all blank, all numbers) — the
        color is likely the only differentiator (e.g. red = overdue, green = on-track).
        This is a POSSIBLE finding requiring manual review.
        """
        colored_cells_no_pattern: List[str] = []

        # Sample first 50 rows to avoid performance issues on large sheets.
        # Track colored blank cells across the sampled area, not only per row.
        try:
            sample_rows = list(ws.iter_rows(min_row=1, max_row=min(50, ws.max_row or 0)))
        except Exception:
            return

        # Look for colored cells with no textual cue.
        for row in sample_rows:
            for cell in row:
                try:
                    fill = cell.fill
                    has_color = (
                        fill is not None
                        and fill.fill_type not in (None, 'none')
                        and fill.fgColor is not None
                        and str(fill.fgColor.rgb) not in ('00000000', 'FFFFFFFF', '00FFFFFF', 'FF000000')
                    )
                    is_blank = cell.value is None or str(cell.value).strip() == ''
                    if has_color and is_blank:
                        cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                        colored_cells_no_pattern.append(cell_ref)
                except Exception:
                    continue

        if len(colored_cells_no_pattern) < 3:
            return

        examples = ', '.join(colored_cells_no_pattern[:8])
        self.fact_sheet.possible_findings.append(Finding(
            criterion_id="1.4.1",
            criterion_name="Use of Color",
            wcag_level="A",
            issue=(
                f"{loc}: Multiple cells appear to use background color as the sole "
                "indicator of meaning (cells are colored but blank)."
            ),
            evidence=f"Colored blank cells found: {examples}.",
            severity=Severity.MODERATE,
            why_it_matters=(
                "Color alone cannot communicate meaning to users who are color-blind or "
                "using high-contrast mode. Status indicators (red/green) must also include "
                "text labels (e.g. 'Yes'/'No', 'Pass'/'Fail', '✓'/'✗')."
            ),
            remediation_steps=[
                f"📍 WHERE TO CHECK: {loc} — cells: {examples}.",
                "  • Add text labels to color-coded cells: 'High', 'Medium', 'Low' alongside the fill color.",
                "  • Use cell text or an adjacent label column rather than relying on fill color alone.",
                "  • Test by viewing sheet in grayscale (Page Layout → Page Color → None) — meaning must still be clear.",
            ],
            confidence_tier=ConfidenceTier.POSSIBLE,
            confidence_label="medium",
            confidence_rationale="Heuristic: colored empty cells detected; manual review required to confirm meaning is color-only.",
            evidence_source=EvidenceSource.XML_INFERRED,
            location=loc,
            remediation_id=f"xlsx_color_only_{ws.title[:20]}",
            remediation_data={"action": "add_text_to_colored_cells", "sheet": ws.title, "cells": colored_cells_no_pattern[:20]},
        ))


    # ── 1.1.1 Images ─────────────────────────────────────────────────────────

    def _rule_1_1_1_images(self, wb) -> None:
        """Check for images embedded in the workbook without alt text."""
        try:
            for ws in wb.worksheets:
                if not hasattr(ws, '_images') or not ws._images:
                    continue
                
                for idx, img in enumerate(ws._images):
                    # openpyxl images have a title/alt text attribute
                    img_alt = getattr(img, 'description', None) or getattr(img, 'name', None) or ''
                    img_alt = (img_alt or '').strip()
                    
                    if not img_alt:
                        self.fact_sheet.confirmed_findings.append(Finding(
                            criterion_id="1.1.1",
                            criterion_name="Non-text Content",
                            wcag_level="A",
                            issue=f"Sheet '{ws.title}': Embedded image #{idx + 1} has no alt text or description.",
                            evidence="Image found without descriptive text or alt attribute.",
                            severity=Severity.SERIOUS,
                            why_it_matters="Screen reader users will not know what the image depicts without alt text.",
                            remediation_steps=[
                                f"📍 WHERE TO FIX: Sheet '{ws.title}', Image #{idx + 1}.",
                                "  • Right-click the image → Alt Text (or Format Picture → Alt Text).",
                                "  • Provide a brief description of what the image shows.",
                                "  • If decorative, use 'Decorative' in the alt text dialog.",
                            ],
                            confidence_tier=ConfidenceTier.CONFIRMED,
                            confidence_label="high",
                            confidence_rationale="Image with no alt text detected in sheet.",
                            evidence_source=EvidenceSource.XML_DIRECT,
                            location=f"Sheet '{ws.title}'",
                            remediation_id=f"xlsx_image_alt_{ws.title[:15]}_{idx}",
                        ))
        except Exception:
            pass

    # ── 1.3.1 Freeform Layout ───────────────────────────────────────────────────

    def _rule_1_3_1_freeform_layout(self, ws, loc: str) -> None:
        """Detect spreadsheets with scattered data (freeform) that lack table structure.
        
        Freeform layouts have large gaps, data in scattered cells, and no clear
        row/column pattern. These are hard for screen reader users to navigate
        because there's no semantic structure to exploit.
        """
        try:
            # Count populated cells and gaps in first 20 rows
            populated_cells = 0
            empty_rows = 0
            max_col_used = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), 1):
                row_has_data = False
                for col_idx, cell in enumerate(row, 1):
                    if cell.value is not None:
                        populated_cells += 1
                        row_has_data = True
                        max_col_used = max(max_col_used, col_idx)
                
                if not row_has_data and row_idx > 2:  # Ignore blank header rows
                    empty_rows += 1
            
            # Heuristic: if ≥ 4 empty rows in first 20, or scattered cells, flag as freeform
            if empty_rows >= 4 and populated_cells < 15 and max_col_used > 3:
                self.fact_sheet.possible_findings.append(Finding(
                    criterion_id="1.3.1",
                    criterion_name="Info and Relationships",
                    wcag_level="A",
                    issue=(
                        f"{loc}: Data appears scattered with multiple blank rows — "
                        "this freeform layout lacks semantic table structure for AT navigation."
                    ),
                    evidence=f"Detected {empty_rows} blank rows and {populated_cells} scattered cells across {max_col_used} columns in sample.",
                    severity=Severity.MODERATE,
                    why_it_matters=(
                        "Screen readers cannot navigate freeform layouts effectively. "
                        "Users cannot use row/column headers or table semantics to understand data relationships."
                    ),
                    remediation_steps=[
                        f"📍 WHERE TO CHECK: {loc}.",
                        "  • Reorganize data into a structured table format with headers in row 1.",
                        "  • Avoid blank rows within data — use a single table with headers and data rows.",
                        "  • If multiple tables exist, create separate formatted tables for each logical grouping.",
                    ],
                    confidence_tier=ConfidenceTier.POSSIBLE,
                    confidence_label="medium",
                    confidence_rationale="Heuristic: multiple blank rows and scattered cells detected; manual review to confirm freeform layout.",
                    evidence_source=EvidenceSource.XML_INFERRED,
                    location=loc,
                    remediation_id=f"xlsx_freeform_{ws.title[:20]}",
                ))
        except Exception:
            pass

    # ── 1.3.2 Reading Order ─────────────────────────────────────────────────────

    def _rule_1_3_2_spreadsheet_sequence(self, ws, loc: str) -> None:
        """Detect blank rows/columns that disrupt logical reading order.
        
        Blank rows and columns should be minimal and purposeful. Many blank rows
        force screen reader users to navigate through silence, losing track of
        where they are in the document.
        """
        try:
            # Evaluate blank-row runs only between real content rows to avoid
            # false positives from trailing worksheet emptiness.
            max_rows_to_check = min(30, max(1, ws.max_row or 1))
            row_has_content = []
            for row_idx in range(1, max_rows_to_check + 1):
                values = next(ws.iter_rows(
                    min_row=row_idx,
                    max_row=row_idx,
                    min_col=1,
                    max_col=26,
                    values_only=True,
                ))
                row_has_content.append(any(value is not None for value in values))

            content_indices = [index for index, has_content in enumerate(row_has_content, start=1) if has_content]
            if len(content_indices) < 2:
                return

            first_content = content_indices[0]
            last_content = content_indices[-1]
            consecutive_blank_rows = 0
            max_consecutive = 0
            for row_idx in range(first_content, last_content + 1):
                if not row_has_content[row_idx - 1]:
                    consecutive_blank_rows += 1
                    max_consecutive = max(max_consecutive, consecutive_blank_rows)
                else:
                    consecutive_blank_rows = 0
            
            if max_consecutive >= 4:
                self.fact_sheet.possible_findings.append(Finding(
                    criterion_id="1.3.2",
                    criterion_name="Meaningful Sequence",
                    wcag_level="A",
                    issue=(
                        f"{loc}: {max_consecutive} consecutive blank rows detected — "
                        "this disrupts logical reading order for screen reader navigation."
                    ),
                    evidence=f"Up to {max_consecutive} consecutive blank rows found in first 30 rows of sheet.",
                    severity=Severity.MINOR,
                    why_it_matters=(
                        "Screen reader users listening to a sheet with many blank rows will hear silence "
                        "and lose track of content position. They cannot quickly navigate over empty space."
                    ),
                    remediation_steps=[
                        f"📍 WHERE TO CHECK: {loc}.",
                        "  • Delete unnecessary blank rows: select rows → Right-click → Delete.",
                        "  • Keep at most 1 blank row between logical sections (if at all).",
                        "  • Use table formatting (Insert → Table) to create clear section breaks instead.",
                    ],
                    confidence_tier=ConfidenceTier.POSSIBLE,
                    confidence_label="medium",
                    confidence_rationale="Excessive blank rows detected; may be intentional formatting or clutter.",
                    evidence_source=EvidenceSource.XML_INFERRED,
                    location=loc,
                    remediation_id=f"xlsx_blank_rows_{ws.title[:20]}",
                ))
        except Exception:
            pass

    # ── 1.4.10 Reflow ───────────────────────────────────────────────────────────

    def _rule_1_4_10_reflow(self, ws, loc: str) -> None:
        """Detect content that would overflow at 200% zoom (or with 2x column widths).
        
        For spreadsheets, reflow is less applicable than for documents, but wide
        content (merged cells, narrow columns with long text) can cause horizontal
        scrolling issues, especially on mobile or zoomed displays.
        """
        try:
            # Check for merged cells spanning many columns (>= 6 columns)
            merged_wide = []
            if ws.merged_cells.ranges:
                for merged_range in ws.merged_cells.ranges:
                    # merged_range is like 'A1:F1'; extract column count
                    min_col = merged_range.min_col
                    max_col = merged_range.max_col
                    col_span = max_col - min_col + 1
                    if col_span >= 6:
                        merged_wide.append(f"{merged_range} (spans {col_span} columns)")
            
            if merged_wide:
                self.fact_sheet.possible_findings.append(Finding(
                    criterion_id="1.4.10",
                    criterion_name="Reflow",
                    wcag_level="AA",
                    issue=(
                        f"{loc}: Wide merged cells or fixed-width content (≥6 columns) may cause "
                        "horizontal scrolling at 200% zoom or on narrow displays."
                    ),
                    evidence=f"Detected merged cells: {merged_wide[:5]}.",
                    severity=Severity.MINOR,
                    why_it_matters=(
                        "Users with low vision zoom to 200% or higher. Content that spans 6+ columns "
                        "at normal zoom will definitely require horizontal scrolling at 200%, breaking "
                        "single-column reading flow on mobile devices."
                    ),
                    remediation_steps=[
                        f"📍 WHERE TO CHECK: {loc}.",
                        "  • Avoid wide merged cells: split headers into smaller cells or use separate rows.",
                        "  • Use column auto-fit (Format → Column → Optimal Width) for long text.",
                        "  • Test at 200% zoom: View → Zoom → 200% — can you read one column at a time?",
                    ],
                    confidence_tier=ConfidenceTier.POSSIBLE,
                    confidence_label="medium",
                    confidence_rationale="Wide merged cells detected; actual reflow impact depends on zoom/display context.",
                    evidence_source=EvidenceSource.XML_INFERRED,
                    location=loc,
                    remediation_id=f"xlsx_reflow_{ws.title[:20]}",
                ))
        except Exception:
            pass

    # ── 2.4.6 Headings and Labels — Header Row Quality ───────────────────────
    def _rule_2_4_6_header_quality(self, ws, loc: str) -> None:
        """WCAG 2.4.6 — Detect generic, blank, or duplicate header row labels."""
        try:
            if ws.max_row < 2 or ws.max_column < 2:
                return  # Empty or single-cell sheet
            # Read first row as putative headers
            headers: List[str] = []
            for col_idx in range(1, min(ws.max_column, 50) + 1):
                cell = ws.cell(row=1, column=col_idx)
                value = cell.value
                if value is None:
                    headers.append("")
                else:
                    headers.append(str(value).strip())

            # Skip if first row appears to be data (no headers at all)
            non_empty = [h for h in headers if h]
            if not non_empty:
                return  # Already covered by 1.3.1 header_row

            # Look for problems
            generic_pattern = re.compile(
                r'^(column\s*\d*|col\s*\d*|field\s*\d*|header\s*\d*|data|value|item|untitled|blank|n/?a)$',
                re.IGNORECASE
            )
            generic_headers = [(i + 1, h) for i, h in enumerate(headers) if h and generic_pattern.match(h)]
            blank_in_middle = []
            # blanks between non-empty headers (data without label)
            for i, h in enumerate(headers):
                if not h:
                    # Is there a non-empty header to left and right?
                    has_left = any(headers[j] for j in range(i))
                    has_right = any(headers[j] for j in range(i + 1, len(headers)))
                    if has_left and has_right:
                        blank_in_middle.append(i + 1)

            # Duplicate headers
            seen: Dict[str, List[int]] = {}
            for i, h in enumerate(headers):
                if h:
                    seen.setdefault(h.lower(), []).append(i + 1)
            duplicates = {h: cols for h, cols in seen.items() if len(cols) > 1}

            issues = []
            if generic_headers:
                samples = ", ".join(f"col {c}: '{t}'" for c, t in generic_headers[:3])
                issues.append(f"{len(generic_headers)} generic header(s) — {samples}")
            if blank_in_middle:
                issues.append(f"{len(blank_in_middle)} blank column header(s) at position(s) {blank_in_middle[:5]}")
            if duplicates:
                samples = "; ".join(f"'{h}' in cols {cols}" for h, cols in list(duplicates.items())[:3])
                issues.append(f"{len(duplicates)} duplicate header(s) — {samples}")

            if not issues:
                return

            self.fact_sheet.possible_findings.append(Finding(
                criterion_id="2.4.6",
                criterion_name="Headings and Labels",
                wcag_level="AA",
                issue=f"Header row in {loc} has quality issues: {'; '.join(issues)}.",
                evidence=f"First-row headers: {headers[:10]}",
                severity=Severity.MINOR,
                why_it_matters=(
                    "Screen reader users rely on column headers to understand cell context. "
                    "Generic, blank, or duplicate headers obscure the data's meaning."
                ),
                remediation_steps=[
                    f"📍 WHERE: {loc}, row 1.",
                    "  • Replace generic headers (Column1, Field2) with descriptive names.",
                    "  • Fill blank header cells that appear between data columns.",
                    "  • Make duplicate headers unique (e.g., 'Q1 Sales', 'Q2 Sales' instead of two 'Sales').",
                ],
                confidence_tier=ConfidenceTier.POSSIBLE,
                confidence_label="medium",
                confidence_rationale="Header text quality measured directly; semantic adequacy needs review.",
                evidence_source=EvidenceSource.TEXT_CONTENT,
                location=loc,
                remediation_id=f"xlsx_header_quality_{ws.title[:20]}",
                remediation_data={"sheet": ws.title, "headers": headers[:20]},
            ))
        except Exception:
            pass

    # ── 3.1.2 Language of Parts ──────────────────────────────────────────────
    def _rule_3_1_2_language_of_parts(self, wb) -> None:
        """WCAG 3.1.2 — Detect runs in shared strings or inline strings with
        explicit xml:lang attributes that differ from the document language."""
        import zipfile as _zipfile
        lang_attrs: List[str] = []
        try:
            with open_safe_zip(self.file_bytes) as zf:
                # Shared strings (when openpyxl uses them)
                for candidate in ('xl/sharedStrings.xml',):
                    try:
                        text = zf.read(candidate).decode('utf-8', errors='replace')
                        lang_attrs.extend(re.findall(r'xml:lang=["\']([^"\']+)["\']', text))
                    except KeyError:
                        pass
                # Inline strings live in worksheet XMLs
                for name in zf.namelist():
                    if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
                        try:
                            text = zf.read(name).decode('utf-8', errors='replace')
                            lang_attrs.extend(re.findall(r'xml:lang=["\']([^"\']+)["\']', text))
                        except KeyError:
                            pass
        except (AttributeError, _zipfile.BadZipFile):
            return

        if not lang_attrs:
            return  # No mixed-language markup at all
        doc_lang = (self.fact_sheet.document_language or '').lower()
        unique_langs = set(l.lower() for l in lang_attrs)
        differing = {l for l in unique_langs if l != doc_lang}
        if not differing:
            return
        self.fact_sheet.possible_findings.append(Finding(
            criterion_id="3.1.2",
            criterion_name="Language of Parts",
            wcag_level="AA",
            issue=(
                f"Workbook contains text runs marked with languages "
                f"{sorted(differing)} differing from document default '{doc_lang or 'unset'}'."
            ),
            evidence=f"xml:lang occurrences in shared strings: {len(lang_attrs)} run(s), languages: {sorted(unique_langs)}.",
            severity=Severity.MINOR,
            why_it_matters=(
                "Foreign-language passages need correct language tags so screen readers use the "
                "right pronunciation engine. Verify each marked run is correctly labelled."
            ),
            remediation_steps=[
                "Open the cells with foreign-language text in Excel.",
                "Verify that the language assigned matches the actual text.",
                "If the document has more than one major language, document this for reviewers.",
            ],
            confidence_tier=ConfidenceTier.POSSIBLE,
            confidence_label="medium",
            confidence_rationale="Language markup detected via shared strings XML; semantic correctness needs human verification.",
            evidence_source=EvidenceSource.XML_DIRECT,
            location="Workbook shared strings",
            remediation_id="xlsx_language_of_parts",
            remediation_data={"languages": sorted(unique_langs), "count": len(lang_attrs)},
        ))

    # ── Phase C: 1.4.3 Cell Text Contrast ───────────────────────────────────
    def _rule_1_4_3_text_contrast(self, ws, loc: str) -> None:
        """WCAG 1.4.3 — Detect cells whose text font color contrasts < 4.5:1
        (or < 3:1 for large/bold text >= 14pt) against the cell's solid fill.

        Strictly deterministic: requires BOTH an explicit RGB font color AND
        an explicit RGB fill color (no theme indices, no defaults). Skips
        empty cells and cells inheriting style from the workbook default.
        """
        from wcag.common.non_text_contrast import normalize_hex
        from wcag.common.utils import contrast_ratio, hex_luminance

        offenders: List[Tuple[str, str, str, float, bool]] = []
        seen_pairs: set = set()
        max_row = min(ws.max_row or 0, 200)
        max_col = min(ws.max_column or 0, 50)
        if max_row == 0 or max_col == 0:
            return

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                # Skip empty cells (no text to evaluate)
                if cell.value is None or str(cell.value).strip() == '':
                    continue
                # Need both an explicit fill and an explicit font color
                fill = cell.fill
                if not fill or fill.patternType != 'solid':
                    continue
                fg_fill = fill.fgColor
                if not fg_fill or fg_fill.type != 'rgb' or not fg_fill.rgb:
                    continue
                fill_hex_raw = str(fg_fill.rgb)
                font = cell.font
                if not font or not font.color or font.color.type != 'rgb' or not font.color.rgb:
                    continue
                font_hex_raw = str(font.color.rgb)

                fill_hex = normalize_hex(fill_hex_raw)
                font_hex = normalize_hex(font_hex_raw)
                if not fill_hex or not font_hex:
                    continue

                # Determine if this is "large text" — Excel default font is 11pt
                size_pt = float(font.size or 11.0)
                is_bold = bool(font.bold)
                large_text = size_pt >= 18.0 or (is_bold and size_pt >= 14.0)
                threshold = 3.0 if large_text else 4.5

                pair_key = (font_hex, fill_hex, large_text)
                if pair_key in seen_pairs:
                    continue
                ratio = contrast_ratio(hex_luminance(font_hex), hex_luminance(fill_hex))
                if ratio >= threshold:
                    continue
                seen_pairs.add(pair_key)
                offenders.append((cell.coordinate, font_hex, fill_hex, ratio, large_text))
                if len(offenders) >= 10:
                    break
            if len(offenders) >= 10:
                break

        if not offenders:
            return
        sample = "; ".join(
            f"{coord} (text #{ft} on fill #{fl}, ratio {r:.2f}:1, "
            f"{'large' if lg else 'normal'})"
            for coord, ft, fl, r, lg in offenders[:3]
        )
        worst_ratio = min(r for _, _, _, r, _ in offenders)
        # Severity escalates if worst ratio is below 3:1 even for normal text
        sev = Severity.SERIOUS if worst_ratio < 3.0 else Severity.MODERATE
        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="1.4.3",
            criterion_name="Contrast (Minimum)",
            wcag_level="AA",
            issue=(
                f"{len(offenders)} cell(s) on '{ws.title}' have text-on-fill contrast below "
                "WCAG threshold (4.5:1 normal, 3:1 large)."
            ),
            evidence=f"Affected cells: {sample}",
            severity=sev,
            why_it_matters=(
                "Low-contrast cell text is hard or impossible to read for users with low vision "
                "or in poor lighting conditions."
            ),
            remediation_steps=[
                "Open the affected cells in Excel and increase contrast between text and fill.",
                "For normal text, ensure ratio >= 4.5:1; for large/bold text (>= 14pt bold or >= 18pt), >= 3:1.",
                "Use Home → Font Color and Fill Color to choose contrasting values.",
                "Verify with WebAIM Contrast Checker (webaim.org/resources/contrastchecker/).",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label=CONFIDENCE_LABEL[EvidenceSource.XML_DIRECT],
            confidence_rationale="Font color and fill color read directly from openpyxl as explicit RGB values.",
            evidence_source=EvidenceSource.XML_DIRECT,
            location=loc,
            remediation_id=f"xlsx_text_contrast_{ws.title}",
        ))

    # ── Phase B: 1.4.11 Non-text Contrast ───────────────────────────────────
    def _rule_1_4_11_non_text_contrast(self, ws, loc: str) -> None:
        """WCAG 1.4.11 — Detect cells with both an explicit border color and a
        solid fill where the contrast between them is below 3:1.

        Strictly deterministic: requires both the border color and the fill
        color to be explicit RGB values (no theme indices, no defaults).
        """
        from wcag.common.non_text_contrast import evaluate_pair, MIN_NON_TEXT_CONTRAST

        offenders: List[Tuple[str, str, str, float]] = []
        seen_pairs = set()
        max_row = min(ws.max_row or 0, 200)  # cap scan for very large sheets
        max_col = min(ws.max_column or 0, 50)
        if max_row == 0 or max_col == 0:
            return

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                fill = cell.fill
                if not fill or fill.patternType != 'solid':
                    continue
                fg = fill.fgColor
                if not fg or fg.type != 'rgb' or not fg.rgb:
                    continue
                fill_hex = str(fg.rgb)
                border = cell.border
                if not border:
                    continue
                # Examine each side; collect first explicit-rgb border color
                border_hex: Optional[str] = None
                for side_name in ('left', 'right', 'top', 'bottom'):
                    side = getattr(border, side_name, None)
                    if side is None or not side.style or side.style == 'none':
                        continue
                    color = side.color
                    if color is None or color.type != 'rgb' or not color.rgb:
                        continue
                    border_hex = str(color.rgb)
                    break
                if not border_hex:
                    continue

                pair_key = (border_hex, fill_hex)
                if pair_key in seen_pairs:
                    continue
                result = evaluate_pair(border_hex, fill_hex)
                if not result:
                    continue
                ratio, ok = result
                if ok:
                    continue
                seen_pairs.add(pair_key)
                offenders.append((cell.coordinate, border_hex, fill_hex, ratio))
                if len(offenders) >= 10:
                    break
            if len(offenders) >= 10:
                break

        if not offenders:
            return
        sample = "; ".join(
            f"{coord} (border #{b[-6:]} on fill #{f[-6:]}, ratio {r:.2f}:1)"
            for coord, b, f, r in offenders[:3]
        )
        self.fact_sheet.confirmed_findings.append(Finding(
            criterion_id="1.4.11",
            criterion_name="Non-text Contrast",
            wcag_level="AA",
            issue=(
                f"{len(offenders)} cell border-on-fill pair(s) on '{ws.title}' have contrast below "
                f"{MIN_NON_TEXT_CONTRAST}:1."
            ),
            evidence=f"Affected cells: {sample}",
            severity=Severity.SERIOUS,
            why_it_matters=(
                "When cell borders define table structure, low contrast between border and fill makes the "
                "structure invisible to low-vision users."
            ),
            remediation_steps=[
                "Increase contrast between border and fill colors to at least 3:1.",
                "Use darker borders on light fills (or vice versa).",
                "Where borders are decorative, consider removing them so structure relies on layout alone.",
            ],
            confidence_tier=ConfidenceTier.CONFIRMED,
            confidence_label=CONFIDENCE_LABEL[EvidenceSource.XML_DIRECT],
            confidence_rationale="Border and fill colors read directly from openpyxl cell style as explicit RGB values.",
            evidence_source=EvidenceSource.XML_DIRECT,
            location=loc,
            remediation_id=f"xlsx_non_text_contrast_{ws.title}",
        ))
