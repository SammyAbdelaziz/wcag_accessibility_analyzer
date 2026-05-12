"""
Generate XLSX test fixture files for WCAG analyzer validation.

Creates various XLSX files that test different WCAG compliance scenarios,
including the 4 new rules: 1.1.1 (images), 1.3.1 (freeform), 1.3.2 (blanks), 1.4.10 (reflow).
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def create_accessible_workbook():
    """Create a fully accessible XLSX workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    # Set workbook properties
    wb.properties.title = "Employee Directory"
    wb.properties.creator = "Test Suite"
    wb.properties.description = "Accessible employee directory"
    
    # Add headers
    headers = ["ID", "Name", "Department", "Salary", "Start Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Add data rows
    data = [
        (1, "Alice Johnson", "Engineering", 85000, "2021-01-15"),
        (2, "Bob Smith", "Sales", 65000, "2019-03-22"),
        (3, "Carol Williams", "HR", 70000, "2020-06-10"),
        (4, "David Brown", "Engineering", 90000, "2018-11-05"),
        (5, "Eve Davis", "Marketing", 72000, "2022-02-14"),
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Set column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    return wb


def create_freeform_layout_workbook():
    """Create a workbook with scattered, freeform layout (WCAG 1.3.1 issue)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scattered"
    
    wb.properties.title = "Freeform Report"
    
    # Scattered data across the sheet
    ws['A1'] = "Q1 Report"
    ws['C3'] = "Revenue"
    ws['D3'] = 150000
    ws['F5'] = "Expenses"
    ws['G5'] = 95000
    ws['B8'] = "Net Profit"
    ws['C8'] = 55000
    ws['E11'] = "Growth Rate"
    ws['F11'] = "12.5%"
    ws['D15'] = "Notes: Positive trend expected in Q2"
    
    return wb


def create_excessive_blank_rows_workbook():
    """Create a workbook with excessive blank rows (WCAG 1.3.2 issue)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ExcessiveBlanks"
    
    wb.properties.title = "Report with Blanks"
    
    # Section 1
    ws['A1'] = "Section 1"
    ws['A2'] = "Data Point A"
    ws['A3'] = "Data Point B"
    
    # 5 blank rows
    for i in range(4, 9):
        pass  # Leave blank
    
    # Section 2
    ws['A9'] = "Section 2"
    ws['A10'] = "Data Point C"
    ws['A11'] = "Data Point D"
    
    # 4 blank rows
    for i in range(12, 16):
        pass  # Leave blank
    
    # Section 3
    ws['A16'] = "Section 3"
    ws['A17'] = "Final Data"
    
    return wb


def create_wide_merged_cells_workbook():
    """Create a workbook with wide merged cells (WCAG 1.4.10 issue)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "WideHeader"
    
    wb.properties.title = "Report with Wide Headers"
    
    # Merge A1:H1 (8 columns - exceeds 6 column threshold)
    ws.merge_cells('A1:H1')
    ws['A1'] = "Quarterly Performance Report - Wide Header"
    
    # Normal data below
    headers = ["Week", "Q1 Sales", "Q2 Sales", "Growth", "Status", "Notes", "Owner", "Date"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=header)
    
    # Sample data
    ws['A3'] = "Week 1"
    ws['B3'] = 10000
    ws['C3'] = 12000
    ws['D3'] = "20%"
    ws['E3'] = "On Track"
    
    return wb


def create_color_only_no_text_workbook():
    """Create a workbook with color used as only indicator (WCAG 1.4.1 issue)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ColorOnly"
    
    wb.properties.title = "Color Status Report"
    
    # Green fill for "Pass"
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    ws['A1'] = ""  # No text, just color
    ws['A1'].fill = green_fill
    
    # Red fill for "Fail"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws['A2'] = ""  # No text, just color
    ws['A2'].fill = red_fill
    
    # Yellow fill for "Pending"
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['A3'] = ""  # No text, just color
    ws['A3'].fill = yellow_fill
    
    return wb


def create_tiny_text_workbook():
    """Create a workbook with tiny text (WCAG 1.4.4 issue)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TinyText"
    
    wb.properties.title = "Report with Tiny Font"
    
    # Normal text
    ws['A1'] = "This is normal size"
    ws['A1'].font = Font(size=12)
    
    # Tiny text (6pt - below 8pt minimum for WCAG AA)
    ws['A2'] = "This is tiny (6pt) - hard to read"
    ws['A2'].font = Font(size=6)
    
    # Also tiny
    ws['A3'] = "This is 7pt - barely visible"
    ws['A3'].font = Font(size=7)
    
    return wb


def create_missing_column_headers_workbook():
    """Create a workbook without column headers (WCAG 1.3.1 issue)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NoHeaders"
    
    wb.properties.title = "Report Without Headers"
    
    # Data without headers - starts directly with values
    ws['A1'] = "Alice"
    ws['B1'] = "Engineering"
    ws['C1'] = 85000
    
    ws['A2'] = "Bob"
    ws['B2'] = "Sales"
    ws['C2'] = 65000
    
    return wb


def create_all_test_files(output_dir: str = None):
    """Generate all test XLSX files."""
    if output_dir is None:
        output_dir = Path(__file__).parent
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    files = {
        "accessible_workbook.xlsx": create_accessible_workbook(),
        "freeform_layout.xlsx": create_freeform_layout_workbook(),
        "excessive_blank_rows.xlsx": create_excessive_blank_rows_workbook(),
        "wide_merged_cells.xlsx": create_wide_merged_cells_workbook(),
        "color_only_no_text.xlsx": create_color_only_no_text_workbook(),
        "tiny_text.xlsx": create_tiny_text_workbook(),
        "missing_headers.xlsx": create_missing_column_headers_workbook(),
    }
    
    for filename, workbook in files.items():
        filepath = output_dir / filename
        workbook.save(str(filepath))
        print(f"✅ Created: {filepath.name}")
    
    print(f"\n📊 Test files created in: {output_dir}")
    return output_dir


if __name__ == "__main__":
    create_all_test_files()
