"""Phase C (XLSX 1.4.3 cell text contrast) + Phase D (PDF 1.3.2 reading order).

Each rule has at least one PASS and one FAIL test that exercises the
strict-deterministic decision logic end-to-end through the analyzer.
"""
import io
import unittest

from wcag.analyzers.xlsx_analyzer import XlsxAnalyzer
from wcag.analyzers.pdf_analyzer import PdfAnalyzer


def _confirmed(fs, criterion: str):
    return [f for f in fs.confirmed_findings if f.criterion_id == criterion]


# ─────────────────────────────────────────────────────────────────────────────
# Phase C — XLSX 1.4.3 Cell Text Contrast
# ─────────────────────────────────────────────────────────────────────────────
def _build_xlsx_with_text_style(font_argb: str, fill_argb: str,
                                size_pt: float = 11.0,
                                bold: bool = False,
                                value: str = "Sample text") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Color

    wb = Workbook()
    ws = wb.active
    ws['A1'] = value
    ws['A1'].font = Font(color=Color(rgb=font_argb), size=size_pt, bold=bold)
    ws['A1'].fill = PatternFill(patternType='solid', fgColor=Color(rgb=fill_argb))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestXlsx143TextContrast(unittest.TestCase):
    def test_pass_black_on_white_normal(self):
        # 21:1 — should NOT trigger
        data = _build_xlsx_with_text_style('FF000000', 'FFFFFFFF')
        fs = XlsxAnalyzer(data, 'pass.xlsx').analyze()
        self.assertEqual(len(_confirmed(fs, "1.4.3")), 0)

    def test_fail_light_gray_on_white_normal(self):
        # #999999 on white = 2.85:1 — fails 4.5:1 threshold for normal text
        data = _build_xlsx_with_text_style('FF999999', 'FFFFFFFF')
        fs = XlsxAnalyzer(data, 'fail.xlsx').analyze()
        self.assertGreater(len(_confirmed(fs, "1.4.3")), 0)

    def test_pass_large_bold_at_3to1(self):
        # #767676 on white = ~4.54:1 normal pass, but use #888 (~3.5:1) with
        # bold 14pt to confirm the relaxed 3:1 threshold for large/bold text.
        data = _build_xlsx_with_text_style(
            'FF888888', 'FFFFFFFF', size_pt=14.0, bold=True
        )
        fs = XlsxAnalyzer(data, 'pass_large.xlsx').analyze()
        self.assertEqual(len(_confirmed(fs, "1.4.3")), 0)

    def test_skip_when_font_color_not_explicit(self):
        # Default font color (no explicit RGB) — strict rule must NOT fire.
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Color
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Default colored text"
        ws['A1'].fill = PatternFill(patternType='solid',
                                    fgColor=Color(rgb='FFCCCCCC'))
        buf = io.BytesIO()
        wb.save(buf)
        fs = XlsxAnalyzer(buf.getvalue(), 'skip.xlsx').analyze()
        self.assertEqual(len(_confirmed(fs, "1.4.3")), 0)

    def test_skip_when_cell_empty(self):
        # Cell with explicit colors but no value — must NOT fire.
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Color
        wb = Workbook()
        ws = wb.active
        ws['A1'].font = Font(color=Color(rgb='FF999999'))
        ws['A1'].fill = PatternFill(patternType='solid',
                                    fgColor=Color(rgb='FFFFFFFF'))
        buf = io.BytesIO()
        wb.save(buf)
        fs = XlsxAnalyzer(buf.getvalue(), 'empty.xlsx').analyze()
        self.assertEqual(len(_confirmed(fs, "1.4.3")), 0)


# ─────────────────────────────────────────────────────────────────────────────
# Phase D — PDF 1.3.2 Meaningful Sequence (sparse structure tree)
# ─────────────────────────────────────────────────────────────────────────────
def _build_pdf_tagged_with_n_struct_elements(n_pages: int, n_elements: int) -> bytes:
    """Build a tagged PDF with ``n_pages`` pages and exactly ``n_elements``
    structure elements under /StructTreeRoot/Document."""
    import pikepdf
    from pikepdf import Pdf, Dictionary, Array, Name

    pdf = Pdf.new()
    for _ in range(n_pages):
        pdf.add_blank_page(page_size=(612, 792))

    # Build a /Document element with N /P children.
    doc_kids = []
    for i in range(n_elements):
        p = pdf.make_indirect(Dictionary(
            Type=Name('/StructElem'),
            S=Name('/P'),
            P=None,  # patched below
        ))
        doc_kids.append(p)

    document = pdf.make_indirect(Dictionary(
        Type=Name('/StructElem'),
        S=Name('/Document'),
        K=Array(doc_kids) if doc_kids else Array([]),
    ))
    for kid in doc_kids:
        kid['/P'] = document

    struct_root = pdf.make_indirect(Dictionary(
        Type=Name('/StructTreeRoot'),
        K=document,
    ))
    document['/P'] = struct_root

    pdf.Root['/StructTreeRoot'] = struct_root
    pdf.Root['/MarkInfo'] = Dictionary(Marked=True)

    buf = io.BytesIO()
    pdf.save(buf)
    return buf.getvalue()


class TestPdf132ReadingOrder(unittest.TestCase):
    def test_pass_dense_structure(self):
        # 3 pages, 6 structure elements (>= page_count) — should NOT fire 1.3.2.
        data = _build_pdf_tagged_with_n_struct_elements(n_pages=3, n_elements=6)
        fs = PdfAnalyzer(data, 'dense.pdf').analyze()
        self.assertEqual(len(_confirmed(fs, "1.3.2")), 0)

    def test_fail_sparse_structure(self):
        # 5 pages, 1 structure element — should fire 1.3.2.
        data = _build_pdf_tagged_with_n_struct_elements(n_pages=5, n_elements=1)
        fs = PdfAnalyzer(data, 'sparse.pdf').analyze()
        findings = _confirmed(fs, "1.3.2")
        self.assertEqual(len(findings), 1)
        self.assertIn("structure element", findings[0].issue.lower())
        self.assertEqual(findings[0].remediation_id, "pdf_reading_order_sparse")

    def test_skip_when_untagged(self):
        # Untagged PDF — 1.3.1 covers it; 1.3.2 must not fire.
        import pikepdf
        pdf = pikepdf.Pdf.new()
        pdf.add_blank_page(page_size=(612, 792))
        pdf.add_blank_page(page_size=(612, 792))
        buf = io.BytesIO()
        pdf.save(buf)
        fs = PdfAnalyzer(buf.getvalue(), 'untagged.pdf').analyze()
        self.assertEqual(len(_confirmed(fs, "1.3.2")), 0)


if __name__ == '__main__':
    unittest.main()
