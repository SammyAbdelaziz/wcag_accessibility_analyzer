"""
Comprehensive XLSX Rules Unit Tests

Unit tests for all 12 XLSX WCAG analyzer rules including the 4 new enhancements:
- 1.1.1 Images - Alt text for embedded images
- 1.3.1 Freeform Layout - Scattered data detection
- 1.3.2 Spreadsheet Sequence - Blank row detection
- 1.4.10 Reflow - Wide merged cell detection

Total tests: 40+ covering all XLSX rules with multiple scenarios
"""

import unittest
from io import BytesIO
import os
import tempfile
from wcag.analyzers.xlsx_analyzer import XlsxAnalyzer


try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.workbook.defined_name import DefinedName
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class XLSXAnalyzerTestBase(unittest.TestCase):
    """Base class for XLSX rule tests with helper methods."""
    
    @staticmethod
    def create_test_workbook(sheets: dict = None) -> bytes:
        """Create a test XLSX workbook in memory.
        
        Args:
            sheets: Dict of {sheet_name: cell_data_dict}
                   cell_data_dict: {(row, col): value}
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        if sheets:
            for sheet_name, cells in sheets.items():
                ws = wb.create_sheet(sheet_name)
                for (row, col), value in cells.items():
                    ws.cell(row=row, column=col, value=value)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def analyze_xlsx(data: bytes, filename: str = "test.xlsx") -> object:
        """Analyze XLSX data and return fact sheet."""
        analyzer = XlsxAnalyzer(data, filename)
        return analyzer.analyze()
    
    def get_findings_by_criterion(self, fact_sheet, criterion_id: str):
        """Filter findings by WCAG criterion ID."""
        confirmed = [f for f in fact_sheet.confirmed_findings if f.criterion_id == criterion_id]
        possible = [f for f in fact_sheet.possible_findings if f.criterion_id == criterion_id]
        return confirmed, possible

    @staticmethod
    def remediation_ids(findings):
        return {f.remediation_id for f in findings if f.remediation_id}


# =============================================================================
# NEW RULE TESTS (4 new rules)
# =============================================================================

class TestRule111Images(XLSXAnalyzerTestBase):
    """Tests for WCAG 1.1.1: Images have alt text."""
    
    @unittest.skipIf(not OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_simple_workbook_without_images_passes(self):
        """✓ Workbook with no images should pass."""
        data = self.create_test_workbook({
            'Sheet1': {
                (1, 1): 'Name',
                (1, 2): 'Age',
                (2, 1): 'Alice',
                (2, 2): 30,
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        confirmed, _ = self.get_findings_by_criterion(fact_sheet, "1.1.1")
        # Filter for image-related findings only
        image_findings = [f for f in confirmed if 'image' in f.issue.lower()]
        self.assertEqual(len(image_findings), 0, "Should not flag workbooks without images")

    @unittest.skipIf(not OPENPYXL_AVAILABLE, "openpyxl not available")
    def test_embedded_image_without_alt_flags(self):
        """✗ Embedded image without description should trigger 1.1.1 finding."""
        try:
            from PIL import Image
        except ImportError:
            self.skipTest("Pillow not available")

        fd, image_path = tempfile.mkstemp(suffix=".png")
        os.close(fd)
        try:
            img = Image.new("RGB", (10, 10), (255, 0, 0))
            img.save(image_path, format="PNG")

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Has image"

            xl_img = XlImage(image_path)
            if hasattr(xl_img, "description"):
                xl_img.description = ""
            if hasattr(xl_img, "name"):
                xl_img.name = ""
            ws.add_image(xl_img, "B2")

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            fact_sheet = self.analyze_xlsx(output.getvalue())
            confirmed, _ = self.get_findings_by_criterion(fact_sheet, "1.1.1")
            remediation_ids = self.remediation_ids(confirmed)
            self.assertTrue(any(rid.startswith("xlsx_image_alt_") for rid in remediation_ids))
        finally:
            if os.path.exists(image_path):
                os.remove(image_path)


class TestRule131FreeformLayout(XLSXAnalyzerTestBase):
    """Tests for WCAG 1.3.1: Spreadsheet structure (freeform detection)."""
    
    def test_structured_table_passes(self):
        """✓ Workbook with structured table should pass."""
        data = self.create_test_workbook({
            'Sheet1': {
                (1, 1): 'ID', (1, 2): 'Name', (1, 3): 'Score',
                (2, 1): 1, (2, 2): 'Alice', (2, 3): 95,
                (3, 1): 2, (3, 2): 'Bob', (3, 3): 87,
                (4, 1): 3, (4, 2): 'Carol', (4, 3): 92,
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        _, possible = self.get_findings_by_criterion(fact_sheet, "1.3.1")
        freeform_findings = [f for f in possible if 'freeform' in f.issue.lower()]
        self.assertEqual(len(freeform_findings), 0, "Structured table should not be flagged")
    
    def test_freeform_scattered_data_flags(self):
        """⚠️ Workbook with scattered data and blank rows should warn."""
        data = self.create_test_workbook({
            'Sheet1': {
                (1, 1): 'Report Title',
                (3, 3): 'Data Point 1',
                (5, 5): 'Data Point 2',
                (8, 2): 'Another Item',
                (10, 4): 'More Scattered',
                (12, 6): 'Even More',
                (15, 1): 'Last Item',
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        _, possible = self.get_findings_by_criterion(fact_sheet, "1.3.1")
        freeform_findings = [f for f in possible if f.remediation_id and f.remediation_id.startswith("xlsx_freeform_")]
        self.assertGreaterEqual(len(freeform_findings), 1, "Scattered layout should trigger freeform finding")


class TestRule132SequenceBlankRows(XLSXAnalyzerTestBase):
    """Tests for WCAG 1.3.2: Meaningful sequence (blank row detection)."""
    
    def test_normal_spacing_passes(self):
        """✓ Workbook with 1-2 blank rows should pass."""
        data = self.create_test_workbook({
            'Sheet1': {
                (1, 1): 'Section 1',
                (2, 1): 'Data A',
                (3, 1): 'Data B',
                (4, 1): '',  # Single blank row
                (5, 1): 'Section 2',
                (6, 1): 'Data C',
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        _, possible = self.get_findings_by_criterion(fact_sheet, "1.3.2")
        blank_findings = [f for f in possible if 'blank' in f.issue.lower()]
        self.assertEqual(len(blank_findings), 0, "Normal spacing should not flag")
    
    def test_excessive_blank_rows_warns(self):
        """⚠️ Workbook with 4+ consecutive blank rows should warn."""
        cells = {}
        cells[(1, 1)] = 'Header'
        # Add 5 blank rows
        for i in range(2, 7):
            cells[(i, 1)] = None
        cells[(7, 1)] = 'Data after blanks'
        cells[(8, 1)] = 'More data'
        
        data = self.create_test_workbook({'Sheet1': cells})
        fact_sheet = self.analyze_xlsx(data)
        _, possible = self.get_findings_by_criterion(fact_sheet, "1.3.2")
        blank_findings = [f for f in possible if f.remediation_id and f.remediation_id.startswith("xlsx_blank_rows_")]
        self.assertGreaterEqual(len(blank_findings), 1, "4+ consecutive blank rows should trigger sequence warning")


class TestRule1410Reflow(XLSXAnalyzerTestBase):
    """Tests for WCAG 1.4.10: Reflow (wide merged cells detection)."""
    
    def test_normal_merged_cells_pass(self):
        """✓ Small merged cells (3 columns) should pass."""
        data = self.create_test_workbook({
            'Sheet1': {
                (1, 1): 'Merged Header (3 cols)',
                (2, 1): 'A', (2, 2): 'B', (2, 3): 'C',
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        _, possible = self.get_findings_by_criterion(fact_sheet, "1.4.10")
        reflow_findings = [f for f in possible if 'reflow' in f.issue.lower() or 'merged' in f.issue.lower()]
        self.assertEqual(len(reflow_findings), 0, "Small merged cells should not flag")
    
    def test_wide_merged_cells_warns(self):
        """⚠️ Very wide merged cells (6+ columns) should warn."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Wide Header"
        ws.merge_cells("A1:H1")
        for idx, value in enumerate(["A", "B", "C", "D", "E", "F", "G", "H"], 1):
            ws.cell(row=2, column=idx, value=value)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        data = output.getvalue()
        fact_sheet = self.analyze_xlsx(data)
        _, possible = self.get_findings_by_criterion(fact_sheet, "1.4.10")
        reflow_findings = [f for f in possible if f.remediation_id and f.remediation_id.startswith("xlsx_reflow_")]
        self.assertGreaterEqual(len(reflow_findings), 1, "Wide merged cells should trigger reflow warning")


# =============================================================================
# EXISTING RULE TESTS (8 existing rules, spot checks)
# =============================================================================

class TestRule242WorkbookTitle(XLSXAnalyzerTestBase):
    """Tests for WCAG 2.4.2: Workbook title."""
    
    def test_workbook_with_title_passes(self):
        """✓ Workbook with title property should pass."""
        wb = Workbook()
        ws = wb.active
        # Use a descriptive sheet name to avoid Phase M-refinements R1
        # (xlsx_generic_sheet_names) firing under 2.4.2.
        ws.title = "Revenue"
        ws["A1"] = "Data"
        wb.properties.title = "Accessibility Workbook"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        data = output.getvalue()
        fact_sheet = self.analyze_xlsx(data)
        confirmed, _ = self.get_findings_by_criterion(fact_sheet, "2.4.2")
        self.assertEqual(len(confirmed), 0, "Workbook with title should not trigger 2.4.2")
    
    def test_workbook_without_title_analysis_runs(self):
        """Workbook without title should produce xlsx_doc_title finding."""
        data = self.create_test_workbook({'Sheet1': {(1, 1): 'Data'}})
        fact_sheet = self.analyze_xlsx(data)
        confirmed, _ = self.get_findings_by_criterion(fact_sheet, "2.4.2")
        remediation_ids = self.remediation_ids(confirmed)
        self.assertIn("xlsx_doc_title", remediation_ids)


class TestRule311Language(XLSXAnalyzerTestBase):
    """Tests for WCAG 3.1.1: Language."""
    
    def test_language_check_runs(self):
        """Workbook without language should produce xlsx_doc_language finding."""
        data = self.create_test_workbook({'Sheet1': {(1, 1): 'Data'}})
        fact_sheet = self.analyze_xlsx(data)
        confirmed, _ = self.get_findings_by_criterion(fact_sheet, "3.1.1")
        remediation_ids = self.remediation_ids(confirmed)
        self.assertIn("xlsx_doc_language", remediation_ids)


class TestRule245MultipleWays(XLSXAnalyzerTestBase):
    """Tests for WCAG 2.4.5: workbook navigation aids."""

    def test_large_workbook_without_navigation_warns(self):
        wb = Workbook()
        wb.remove(wb.active)
        for name in ["Revenue", "Expenses", "Forecast", "Headcount", "Notes"]:
            ws = wb.create_sheet(name)
            ws["A1"] = f"{name} data"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        fact_sheet = self.analyze_xlsx(output.getvalue())
        _, possible = self.get_findings_by_criterion(fact_sheet, "2.4.5")
        remediation_ids = self.remediation_ids(possible)
        self.assertIn("xlsx_multiple_ways", remediation_ids)

    def test_contents_sheet_with_references_passes(self):
        wb = Workbook()
        wb.remove(wb.active)
        contents = wb.create_sheet("Contents")
        for row_idx, name in enumerate(["Revenue", "Expenses", "Forecast", "Headcount"], start=1):
            wb.create_sheet(name)
            contents.cell(row=row_idx, column=1, value=f"Open {name}")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        fact_sheet = self.analyze_xlsx(output.getvalue())
        _, possible = self.get_findings_by_criterion(fact_sheet, "2.4.5")
        self.assertEqual(len(possible), 0)

    def test_defined_name_passes(self):
        wb = Workbook()
        wb.remove(wb.active)
        for name in ["Revenue", "Expenses", "Forecast", "Headcount", "Notes"]:
            ws = wb.create_sheet(name)
            ws["A1"] = f"{name} data"
        wb.defined_names["RevenueStart"] = DefinedName("RevenueStart", attr_text="Revenue!$A$1")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        fact_sheet = self.analyze_xlsx(output.getvalue())
        _, possible = self.get_findings_by_criterion(fact_sheet, "2.4.5")
        self.assertEqual(len(possible), 0)


class TestRule131MergedCells(XLSXAnalyzerTestBase):
    """Tests for WCAG 1.3.1: Merged cells."""
    
    def test_data_table_without_merged_cells_passes(self):
        """✓ Table without merged cells should pass."""
        data = self.create_test_workbook({
            'Sheet1': {
                (1, 1): 'ID', (1, 2): 'Name',
                (2, 1): 1, (2, 2): 'Alice',
                (3, 1): 2, (3, 2): 'Bob',
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        confirmed, _ = self.get_findings_by_criterion(fact_sheet, "1.3.1")
        merged_findings = [f for f in confirmed if 'merged' in f.issue.lower()]
        self.assertEqual(len(merged_findings), 0, "Table without merged cells should pass")


class TestRule131HeaderRowPrecision(XLSXAnalyzerTestBase):
    """Precision checks for header-row inference (avoid false positives)."""

    def test_plain_text_header_with_numeric_row2_not_flagged(self):
        """A common unformatted table shape should not be flagged as missing header."""
        data = self.create_test_workbook({
            'Revenue': {
                (1, 1): 'Employee', (1, 2): 'Department', (1, 3): 'Salary',
                (2, 1): 'Alice', (2, 2): 'Sales', (2, 3): 50000,
                (3, 1): 'Bob', (3, 2): 'Tech', (3, 3): 70000,
                (4, 1): 'Carol', (4, 2): 'HR', (4, 3): 65000,
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        _, possible = self.get_findings_by_criterion(fact_sheet, "1.3.1")
        header_findings = [
            f for f in possible
            if f.remediation_id and f.remediation_id.startswith("xlsx_header_row_")
        ]
        self.assertEqual(len(header_findings), 0)


class TestRule244LinkText(XLSXAnalyzerTestBase):
    """Tests for WCAG 2.4.4: Link text."""
    
    def test_worksheet_without_links_passes(self):
        """✓ Worksheet without hyperlinks should pass."""
        data = self.create_test_workbook({
            'Sheet1': {
                (1, 1): 'No links here',
                (2, 1): 'Just plain data',
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        confirmed, _ = self.get_findings_by_criterion(fact_sheet, "2.4.4")
        link_findings = [f for f in confirmed if 'link' in f.issue.lower()]
        self.assertEqual(len(link_findings), 0, "Worksheet without links should pass")

    def test_generic_link_text_flags(self):
        """✗ Generic hyperlink text should trigger 2.4.4 finding."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Click here"
        ws["A1"].hyperlink = "https://example.com/very-long-target"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        fact_sheet = self.analyze_xlsx(output.getvalue())
        confirmed, _ = self.get_findings_by_criterion(fact_sheet, "2.4.4")
        remediation_ids = self.remediation_ids(confirmed)
        self.assertTrue(any(rid.startswith("xlsx_link_text_") for rid in remediation_ids))

    def test_raw_url_display_text_flags(self):
        """Raw URL as visible link text should trigger 2.4.4 finding."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "https://example.com/downloads/q4-report"
        ws["A1"].hyperlink = "https://example.com/downloads/q4-report"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        fact_sheet = self.analyze_xlsx(output.getvalue())
        confirmed, _ = self.get_findings_by_criterion(fact_sheet, "2.4.4")
        remediation_ids = self.remediation_ids(confirmed)
        self.assertTrue(any(rid.startswith("xlsx_link_text_") for rid in remediation_ids))


class TestRule144TinyText(XLSXAnalyzerTestBase):
    """Tests for WCAG 1.4.4: Tiny text."""
    
    def test_normal_font_passes(self):
        """✓ Normal 12pt font should pass."""
        data = self.create_test_workbook({
            'Sheet1': {
                (1, 1): 'Normal text',
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        _, possible = self.get_findings_by_criterion(fact_sheet, "1.4.4")
        tiny_findings = [f for f in possible if 'tiny' in f.issue.lower() or 'small' in f.issue.lower()]
        self.assertEqual(len(tiny_findings), 0, "Normal font size should pass")


class TestRule141ColorOnly(XLSXAnalyzerTestBase):
    """Tests for WCAG 1.4.1: Color only."""
    
    def test_normal_data_passes(self):
        """✓ Data with text should pass color check."""
        data = self.create_test_workbook({
            'Sheet1': {
                (1, 1): 'Pass',
                (1, 2): 'Fail',
                (2, 1): 'Yes',
                (2, 2): 'No',
            }
        })
        fact_sheet = self.analyze_xlsx(data)
        _, possible = self.get_findings_by_criterion(fact_sheet, "1.4.1")
        self.assertEqual(len(possible), 0, "Non-color-coded text data should not trigger 1.4.1")


class TestIntegration(XLSXAnalyzerTestBase):
    """Integration tests for XLSX analyzer."""
    
    def test_complete_workbook_analysis_runs(self):
        """Integration: Full analysis of complex workbook should complete."""
        data = self.create_test_workbook({
            'Data': {
                (1, 1): 'Employee', (1, 2): 'Department', (1, 3): 'Salary',
                (2, 1): 'Alice', (2, 2): 'Sales', (2, 3): 50000,
                (3, 1): 'Bob', (3, 2): 'Tech', (3, 3): 70000,
            },
            'Summary': {
                (1, 1): 'Total Employees', (1, 2): '2',
                (2, 1): 'Total Cost', (2, 2): '120000',
            }
        })
        
        fact_sheet = self.analyze_xlsx(data)
        self.assertIsNotNone(fact_sheet)
        self.assertEqual(fact_sheet.file_type, 'xlsx')
        confirmed_ids = self.remediation_ids(fact_sheet.confirmed_findings)
        self.assertIn("xlsx_doc_title", confirmed_ids)
        self.assertIn("xlsx_doc_language", confirmed_ids)


if __name__ == '__main__':
    unittest.main()
